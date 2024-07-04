import re
import openpyxl
from typing import Callable


class Correspondences:
    def __init__(self, name, sheet):
        self._name = name
        self._rules = {sheet.cell(j, 1).value: sheet.cell(j, 2).value for j in range(2, len(sheet['A']))}
        self._rules.pop(None)
        self._worker: Callable[[dict[str, str], str], str] = self._standard_worker

    @property
    def name(self):
        return self._name

    def find_answer(self, value):
        return self._worker(self._rules, value)

    def set_new_worker(self, worker: Callable[[dict[str, str], str], str]):
        self._worker = worker

    def reset_worker(self):
        self._worker = self._standard_worker

    @staticmethod
    def _standard_worker(rules: dict[str, str], value: str):
        value = value.split('>', 1)[1].strip()
        for i in rules.keys():
            if re.search(i, value) is not None:
                return rules[i]
        return None

    @staticmethod
    def get_correspondences(file):
        book = openpyxl.load_workbook(file)
        data = {i: Correspondences(i, book[i]) for i in book.sheetnames}
        book.close()
        return data
